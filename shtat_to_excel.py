from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Side
from connection import Connection


class Shtat_To_Excel:
    def __init__(self, file):
        self.file = file

    def shtat_to_excel(self):
        con = Connection.connect()
        cur = con.cursor()
        cur.execute("SELECT * from salaries WHERE fio NOT LIKE '%Вакансия%' ORDER BY department_code")

        work_data = cur.fetchall()

        # определяем стили
        fill = PatternFill(fill_type='solid',
                           start_color='c1c1c1',
                           end_color='c2c2c2')

        border = Border(left=Side(border_style='thin',
                                  color='FF000000'),
                        right=Side(border_style='thin',
                                   color='FF000000'),
                        top=Side(border_style='thin',
                                 color='FF000000'),
                        bottom=Side(border_style='thin',
                                    color='FF000000'),
                        diagonal=Side(border_style='thin',
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style='thin',
                                     color='FF000000'),
                        vertical=Side(border_style='thin',
                                      color='FF000000'),
                        horizontal=Side(border_style='thin',
                                        color='FF000000')
                        )
        align_center = Alignment(horizontal='center',
                                 vertical='center',
                                 text_rotation=0,
                                 wrap_text=False,
                                 shrink_to_fit=False,
                                 indent=0)
        align_left = Alignment(horizontal='right',
                               vertical='center',
                               text_rotation=0,
                               wrap_text=False,
                               shrink_to_fit=False,
                               indent=0)

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active

        # название страницы
        ws.title = 'Штатное расписание'
        # данные для строк

        rows = [["№ позиции",
                 "Подразделение",
                 "Должность", "ФИО", "Больничные дни", "Отпуск", "Простой"]]

        for i in work_data:
            rows.append([i[0], i[1], i[2], i[6], 0, 0, 0])

        # циклом записываем данные
        for row in rows:
            ws.append(row)

        # раскрашивание фона для заголовков
        for i in range(0, ws.max_column):
            ws[f'{chr(65 + i)}1'].fill = fill

        # увеличиваем все строки по высоте
        max_row = ws.max_row
        i = 1
        while i <= max_row:
            rd = ws.row_dimensions[i]
            rd.height = 16
            i += 1

        # сетка + выравнивание
        for cellObj in ws[f'A1:{chr(64 + ws.max_column)}{ws.max_row}']:
            for cell in cellObj:
                ws[cell.coordinate].border = border
                ws[cell.coordinate].alignment = align_center

        # выравнивание столбца
        for cellObj in ws[f'A1:{chr(64 + ws.max_column)}{ws.max_row}']:
            for cell in cellObj:
                ws[cell.coordinate].alignment = align_left

        # сохранение файла в текущую директорию
        wb.save(self.file)


if __name__ == '__main__':
    save_shtat = Shtat_To_Excel("Штатное.xlsx")
    save_shtat.shtat_to_excel()
