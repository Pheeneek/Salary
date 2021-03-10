from PyQt5 import QtWidgets
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Side
from counting.personel_summary import Personel_summary


class Personel_To_Excel:
    """
    Класс для выгрузки таблицы со штатной численности в файл Excel
    """

    def __init__(self, file: str, qtgui: any) -> None:
        """
        Метод инициализации класса
        :param file: файл для выгрузки
        :param qtgui: экземпляр класса основного окна
        """
        self.gui = qtgui
        self.file = file
        self.work_data = Personel_summary()
        self.data = self.work_data.data_counting()
        self.rows = []
        self.fill = PatternFill(fill_type='solid',
                                start_color='c1c1c1',
                                end_color='c2c2c2')

        self.border = Border(left=Side(border_style='thin',
                                       color='FF000000'),
                             right=Side(border_style='thin',
                                        color='FF000000'),
                             top=Side(border_style='thin',
                                      color='FF000000'),
                             bottom=Side(border_style='thin',
                                         color='FF000000'),
                             diagonal=Side(border_style='thin',
                                           color='FF000000'),
                             diagonal_direction=0,
                             outline=Side(border_style='thin',
                                          color='FF000000'),
                             vertical=Side(border_style='thin',
                                           color='FF000000'),
                             horizontal=Side(border_style='thin',
                                             color='FF000000')
                             )
        self.align_center = Alignment(horizontal='center',
                                      vertical='center',
                                      text_rotation=0,
                                      wrap_text=False,
                                      shrink_to_fit=False,
                                      indent=0)
        self.align_left = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Численность'

    def personel_to_excel(self) -> None:
        """
        Метод выгрузки численности в файл
        :return: None
        """
        self.ws.cell(row=4, column=1)
        self.rows.append(["Подразделения",
                          "Код",
                          "Производственный персонал", "", "",
                          "ПП Всего", "Непроизводственный персонал", "", "НП Всего", "ППП"])
        self.rows.append(["", "", "Сдельщики", "На окладах", "Технологи",
                          "Всего", "Специалисты", "Вспомогательные рабочие",
                          "Всего"])
        current_row = 7
        for i in self.data:
            self.rows.append(["", i, self.data[i]['pp']['sdelka'],
                              self.data[i]['pp']['vremya'],
                              self.data[i]['pp']['tech'],
                              self.data[i]['pp']['summary'],
                              self.data[i]['np']['specialist'],
                              self.data[i]['np']['worker'],
                              self.data[i]['np']['summary'],
                              f"=F{current_row}+I{current_row}"])
            current_row += 1
        current_row -= 1
        self.rows.append(["", "", f"=SUM(C7:C{current_row})",
                          f"=SUM(D7:D{current_row})",
                          f"=SUM(E7:E{current_row})",
                          f"=SUM(F7:F{current_row})",
                          f"=SUM(G7:G{current_row})",
                          f"=SUM(H7:H{current_row})",
                          f"=SUM(I7:I{current_row})",
                          f"=SUM(J7:J{current_row})"])
        for row in self.rows:
            self.ws.append(row)

        for i in range(0, self.ws.max_column):
            self.ws[f'{chr(65 + i)}5'].fill = self.fill

        for i in range(0, self.ws.max_column):
            self.ws[f'{chr(65 + i)}6'].fill = self.fill

        max_row = self.ws.max_row
        i = 1
        while i <= max_row:
            rd = self.ws.row_dimensions[i]
            rd.height = 16
            i += 1

        for cellObj in self.ws[f'A5:{chr(64 + self.ws.max_column)}{self.ws.max_row}']:
            for cell in cellObj:
                self.ws[cell.coordinate].border = self.border
                self.ws[cell.coordinate].alignment = self.align_center

        for cellObj in self.ws[f'A4:{chr(64 + self.ws.max_column)}{self.ws.max_row}']:
            for cell in cellObj:
                self.ws[cell.coordinate].alignment = self.align_left
        self.ws.merge_cells("C5:E5")
        self.ws.merge_cells("G5:H5")
        self.ws.merge_cells("A5:A6")
        self.ws.merge_cells("B5:B6")
        self.ws.merge_cells("F5:F6")
        self.ws.merge_cells("I5:I6")
        self.ws.merge_cells("J5:J6")

        try:
            self.wb.save(self.file)
            self.save_confirmation("Файл отклонений успешно сохранен", "Успех!")
        except PermissionError:
            self.save_confirmation("Не удалось сохранить файл!", "Ошибка!")

    def save_confirmation(self, text: str, title: str) -> None:
        """
        Метод, запускающий информационное окно c результатом записи файла
        :param text: текст сообщения
        :param title: Название окна
        :return: None
        """
        msg = QtWidgets.QMessageBox(self.gui.window)
        msg.setIcon(QtWidgets.QMessageBox.Information)
        msg.setText(text)
        msg.setWindowTitle(title)
        msg.exec()
