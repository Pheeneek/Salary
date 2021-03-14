"""
Файл с классом, реализующим выгрузку таблицы с расчетом стимула в файл Excel
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Side


class StimulToExcel:
    """
     Класс, выгружающий таблицу с расчетом стимула в файл Excel для заполнения отклонений
     """

    def __init__(self, gui, stimul_data, file):
        self.gui = gui
        self.stimul_data = stimul_data
        self.file = file
        self.rows = []
        self.fill = PatternFill(fill_type='solid',
                                start_color='c1c1c1',
                                end_color='c2c2c2')

        self.border = Border(left=Side(border_style='thin',
                                       color='FF000000'),
                             right=Side(border_style='thin',
                                        color='FF000000'),
                             top=Side(border_style='thin',
                                      color='FF000000'),
                             bottom=Side(border_style='thin',
                                         color='FF000000'),
                             diagonal=Side(border_style='thin',
                                           color='FF000000'),
                             diagonal_direction=0,
                             outline=Side(border_style='thin',
                                          color='FF000000'),
                             vertical=Side(border_style='thin',
                                           color='FF000000'),
                             horizontal=Side(border_style='thin',
                                             color='FF000000')
                             )
        self.align_center = Alignment(horizontal='center',
                                      vertical='center',
                                      text_rotation=0,
                                      wrap_text=False,
                                      shrink_to_fit=False,
                                      indent=0)
        self.align_left = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Распределение стимула'

    def stimul_to_excel(self) -> None:
        """
        Метод выгрузки стимула в файл
        :return: None
        """
        self.ws.cell(row=4, column=1)
        self.rows.append(['Подразделение', 'ПП оклады', 'ПП стимул',
                          'НП оклады', 'НП стимул',
                          'Технологи оклады', 'Технологи стимул', 'Процент'])
        for i in self.stimul_data:
            self.rows.append(i)

        for row in self.rows:
            self.ws.append(row)

        for i in range(0, self.ws.max_column):
            self.ws[f'{chr(65 + i)}5'].fill = self.fill

        for i in range(0, self.ws.max_column):
            self.ws[f'{chr(65 + i)}6'].fill = self.fill

        max_row = self.ws.max_row
        i = 1
        while i <= max_row:
            rd = self.ws.row_dimensions[i]
            rd.height = 16
            i += 1

        for cellObj in self.ws[f'A5:{chr(64 + self.ws.max_column)}{self.ws.max_row}']:
            for cell in cellObj:
                self.ws[cell.coordinate].border = self.border
                self.ws[cell.coordinate].alignment = self.align_center

        for cellObj in self.ws[f'A4:{chr(64 + self.ws.max_column)}{self.ws.max_row}']:
            for cell in cellObj:
                self.ws[cell.coordinate].alignment = self.align_left

        try:
            self.wb.save(self.file)
            self.gui.actions.save_confirmation("Файл стимула успешно сохранен", "Успех!")
        except PermissionError:
            self.gui.actions.save_confirmation("Не удалось сохранить файл!", "Ошибка!")
