from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Side
from connection.connection import Connection


class Shtat_To_Excel:
    """
    Класс, выгружающий штатное расписание в файл Excel для заполнения отклонений
    """
    def __init__(self, file: str):
        """
        Метод инициализации класса
        :param file: Название файла для отклонений
        """
        self.file = file
        self.con, self.cur = Connection.connect()
        self.cur.execute("SELECT * from salaries WHERE fio NOT LIKE '%Вакансия%' ORDER BY department_code")
        self.work_data = self.cur.fetchall()

        self.fill = PatternFill(fill_type='solid',
                                start_color='c1c1c1',
                                end_color='c2c2c2')

        self.border = Border(left=Side(border_style='thin',
                                       color='FF000000'),
                             right=Side(border_style='thin',
                                        color='FF000000'),
                             top=Side(border_style='thin',
                                      color='FF000000'),
                             bottom=Side(border_style='thin',
                                         color='FF000000'),
                             diagonal=Side(border_style='thin',
                                           color='FF000000'),
                             diagonal_direction=0,
                             outline=Side(border_style='thin',
                                          color='FF000000'),
                             vertical=Side(border_style='thin',
                                           color='FF000000'),
                             horizontal=Side(border_style='thin',
                                             color='FF000000')
                             )
        self.align_center = Alignment(horizontal='center',
                                      vertical='center',
                                      text_rotation=0,
                                      wrap_text=False,
                                      shrink_to_fit=False,
                                      indent=0)
        self.align_left = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Штатное расписание'

    def shtat_to_excel(self) -> None:
        """
        Метод выгрузки штатного расписания в файл
        :return: None
        """
        rows = [["№ позиции",
                 "Подразделение",
                 "Должность", "ФИО", "Больничные дни", "Отпуск", "Простой"]]

        for i in self.work_data:
            rows.append([i[0], i[1], i[2], i[6], 0, 0, 0])

        for row in rows:
            self.ws.append(row)

        for i in range(0, self.ws.max_column):
            self.ws[f'{chr(65 + i)}1'].fill = self.fill

        max_row = self.ws.max_row
        i = 1
        while i <= max_row:
            rd = self.ws.row_dimensions[i]
            rd.height = 16
            i += 1

        for cellObj in self.ws[f'A1:{chr(64 + self.ws.max_column)}{self.ws.max_row}']:
            for cell in cellObj:
                self.ws[cell.coordinate].border = self.border
                self.ws[cell.coordinate].alignment = self.align_center
        for cellObj in self.ws[f'A1:{chr(64 + self.ws.max_column)}{self.ws.max_row}']:
            for cell in cellObj:
                self.ws[cell.coordinate].alignment = self.align_left

        self.wb.save(self.file)
