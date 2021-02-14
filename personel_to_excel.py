from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Protection, Font, Side
from personel_summary import Personel_summary
from pprint import pprint

work_data = Personel_summary()
data = work_data.data_counting()

# определяем стили
font = Font(name='Arial',
            size=12,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')

fill = PatternFill(fill_type='solid',
                   start_color='c1c1c1',
                   end_color='c2c2c2')

border = Border(left=Side(border_style='thin',
                          color='FF000000'),
                right=Side(border_style='thin',
                           color='FF000000'),
                top=Side(border_style='thin',
                         color='FF000000'),
                bottom=Side(border_style='thin',
                            color='FF000000'),
                diagonal=Side(border_style='thin',
                              color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style='thin',
                             color='FF000000'),
                vertical=Side(border_style='thin',
                              color='FF000000'),
                horizontal=Side(border_style='thin',
                                color='FF000000')
                )
align_center = Alignment(horizontal='center',
                         vertical='center',
                         text_rotation=0,
                         wrap_text=False,
                         shrink_to_fit=False,
                         indent=0)
align_left = Alignment(horizontal='right',
                       vertical='center',
                       text_rotation=0,
                       wrap_text=False,
                       shrink_to_fit=False,
                       indent=0)
number_format = 'General'
protection = Protection(locked=False,
                        hidden=False)

# объект
wb = Workbook()

# активный лист
ws = wb.active

# название страницы
ws.title = 'Численность'
# данные для строк


rows = []
ws.cell(row=4, column=1)
rows.append(["Подразделения",
             "Код",
             "Производственный персонал", "", "",
             "ПП Всего", "Непроизводственный персонал", "", "НП Всего", "ППП"])
rows.append(["", "", "Сдельщики", "На окладах", "Технологи",
             "Всего", "Специалисты", "Вспомогательные рабочие",
             "Всего"])

for i in data:
    rows.append(["", i, data[i]['pp']['sdelka'],
                 data[i]['pp']['vremya'],
                 data[i]['pp']['tech'],
                 data[i]['pp']['summary'],
                 data[i]['np']['specialist'],
                 data[i]['np']['worker'],
                 data[i]['np']['summary'],
                 ])
pprint(rows)

# циклом записываем данные
for row in rows:
    ws.append(row)

# раскрвшивание фона для заголовков
for i in range(0, ws.max_column):
    ws[f'{chr(65 + i)}5'].fill = fill

for i in range(0, ws.max_column):
    ws[f'{chr(65 + i)}6'].fill = fill

# увеличиваем все строки по высоте
max_row = ws.max_row
i = 1
while i <= max_row:
    rd = ws.row_dimensions[i]
    rd.height = 16
    i += 1

# сетка + выравнивание
for cellObj in ws[f'A5:{chr(64 + ws.max_column)}{ws.max_row}']:
    for cell in cellObj:
        ws[cell.coordinate].border = border
        ws[cell.coordinate].alignment = align_center

# выравнивание столбца
for cellObj in ws[f'A4:{chr(64 + ws.max_column)}{ws.max_row}']:
    for cell in cellObj:
        ws[cell.coordinate].alignment = align_left
ws.merge_cells("C5:E5")
ws.merge_cells("G5:H5")
ws.merge_cells("A5:A6")
ws.merge_cells("B5:B6")
ws.merge_cells("F5:F6")
ws.merge_cells("I5:I6")
ws.merge_cells("J5:J6")
# сохранение файла в текущую директорию
wb.save("Численность.xlsx")
